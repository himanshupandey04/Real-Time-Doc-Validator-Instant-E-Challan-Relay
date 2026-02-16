import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- Configuration ---
INPUT_CSV = r"D:\indian_vehicle_dataset.csv"
OUTPUT_XLSX = r"D:\indian_vehicle_dataset_non_ev.xlsx"

# Pure EV manufacturers to remove
EV_MAKES = {"Ather", "Ola", "Okinawa", "Hero Electric", "Pure EV", "Ampere", "Revolt"}

# --- Read CSV and filter ---
with open(INPUT_CSV, "r", encoding="utf-8") as f:
    reader = csv.reader(f)
    header = next(reader)

    # Find column indices
    vehicle_class_idx = header.index("Vehicle_Class")
    fuel_type_idx = header.index("Fuel_Type")
    make_idx = header.index("Make")
    engine_capacity_idx = header.index("Engine_Capacity_CC")

    filtered_rows = []
    removed_count = 0

    for row in reader:
        vehicle_class = row[vehicle_class_idx].strip()
        fuel_type = row[fuel_type_idx].strip()
        make = row[make_idx].strip()

        # Remove if: Electric Vehicle class, Electric fuel, or pure-EV manufacturer
        is_ev = (
            vehicle_class == "Electric Vehicle" or
            fuel_type == "Electric" or
            make in EV_MAKES
        )

        if is_ev:
            removed_count += 1
            continue

        # Fix Engine_Capacity_CC: if 0 for non-EV, it's likely data error — keep as-is
        filtered_rows.append(row)

total_kept = len(filtered_rows)
print(f"Original records: {removed_count + total_kept}")
print(f"Removed (EV): {removed_count}")
print(f"Kept (non-EV): {total_kept}")

# --- Create formatted Excel workbook ---
wb = Workbook()
ws = wb.active
ws.title = "Vehicle Data"

# Styles
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_font = Font(name="Calibri", size=10)
cell_alignment = Alignment(vertical="center", wrap_text=False)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Write header row
for col_idx, col_name in enumerate(header, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data rows
for row_idx, row_data in enumerate(filtered_rows, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = thin_border

# Auto-fit column widths (approximate)
for col_idx in range(1, len(header) + 1):
    col_letter = get_column_letter(col_idx)
    max_length = len(str(header[col_idx - 1]))
    # Sample first 100 data rows for width
    for row_idx in range(2, min(102, total_kept + 2)):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        if cell_value:
            max_length = max(max_length, len(str(cell_value)))
    ws.column_dimensions[col_letter].width = min(max_length + 3, 35)

# Freeze top row for easy scrolling
ws.freeze_panes = "A2"

# Add auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

# Save
wb.save(OUTPUT_XLSX)
print(f"\nExcel file saved: {OUTPUT_XLSX}")
print("Features: formatted headers, auto-column-width, frozen header row, auto-filter enabled")
